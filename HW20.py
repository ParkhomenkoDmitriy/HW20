import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

csv_file_path = "D:/people_with_phones.csv"
data = []

with open(csv_file_path, 'r', encoding='utf-8-sig') as csv_file:
    csv_reader = csv.reader(csv_file)
    for row in csv_reader:
        data.append(row)

format_data = list(zip(*data))

format_data = [row for row in format_data if "Возраст" not in row]

wb = Workbook()
ws = wb.active

for row_index, row in enumerate(format_data):
    for col_index, value in enumerate(row):
        cell = get_column_letter(col_index + 1) + str(row_index + 1)
        ws[cell] = value

excel_file_path = "D:/people_format_no_age.xlsx"
wb.save(excel_file_path)

print(f"Excel-файл сохранен: {excel_file_path}")
